"""
Ingestion de documents : chargement, découpage, ajout au vector store (Chroma via Haystack).
"""

from pathlib import Path

from langchain_community.document_loaders import PyPDFLoader, TextLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document as LangChainDocument
from openpyxl import load_workbook
from haystack import Document as HaystackDocument

from app.config import get_settings
from app.haystack_rag import index_documents_haystack


def _load_xlsx(file_path: str) -> list[LangChainDocument]:
    """Charge un fichier Excel .xlsx et retourne un Document LangChain par feuille."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))
        text = "\n".join(rows).strip()
        if text:
            docs.append(
                LangChainDocument(
                    page_content=text,
                    metadata={"source_file": Path(file_path).name, "sheet": sheet_name},
                )
            )
    wb.close()
    return docs


def get_loader_for_path(file_path: str):
    """Retourne le loader LangChain adapté au type de fichier (ou None pour xlsx, géré à part)."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return PyPDFLoader(file_path)
    if suffix in (".txt", ".md"):
        return TextLoader(file_path, encoding="utf-8")
    if suffix == ".xlsx":
        return None  # géré par load_and_split_documents
    raise ValueError(f"Type de fichier non supporté : {suffix}")


def load_and_split_documents(file_path: str) -> list[LangChainDocument]:
    """Charge un fichier et le découpe en chunks (documents LangChain)."""
    settings = get_settings()
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        docs = _load_xlsx(file_path)
    else:
        loader = get_loader_for_path(file_path)
        docs = loader.load()
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        length_function=len,
    )
    return splitter.split_documents(docs)


def _lc_to_haystack_docs(lc_docs: list[LangChainDocument], file_path: str, metadata: dict | None = None) -> list[HaystackDocument]:
    """Convertit des documents LangChain en documents Haystack."""
    out = []
    for d in lc_docs:
        meta = dict(d.metadata or {})
        meta["source_file"] = str(Path(file_path).name)
        if metadata:
            meta.update(metadata)
        out.append(HaystackDocument(content=d.page_content, meta=meta))
    return out


def ingest_file(file_path: str, metadata: dict | None = None) -> list[str]:
    """
    Ingère un fichier dans Chroma via Haystack (embedding + écriture).
    Retourne une liste d'ids factices pour compatibilité API (count = len(ids)).
    """
    docs = load_and_split_documents(file_path)
    if not docs:
        return []
    for d in docs:
        d.metadata = d.metadata or {}
        d.metadata["source_file"] = str(Path(file_path).name)
        if metadata:
            d.metadata.update(metadata)
    haystack_docs = _lc_to_haystack_docs(docs, file_path, metadata)
    count = index_documents_haystack(haystack_docs)
    return [str(i) for i in range(count)]


def ingest_bytes(content: bytes, filename: str, metadata: dict | None = None) -> list[str]:
    """
    Ingère du contenu binaire (upload) dans Chroma.
    Écrit temporairement sur disque pour les loaders qui lisent des fichiers.
    """
    import tempfile
    suffix = Path(filename).suffix.lower()
    if suffix not in (".pdf", ".txt", ".md", ".xlsx"):
        raise ValueError(f"Type non supporté : {suffix}. Utilisez .pdf, .txt, .md ou .xlsx")
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
        f.write(content)
        path = f.name
    try:
        return ingest_file(path, metadata=metadata or {"filename": filename})
    finally:
        Path(path).unlink(missing_ok=True)
